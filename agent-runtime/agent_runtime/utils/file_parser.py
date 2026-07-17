# -*- coding: UTF-8 -*-
# Copyright (c) Huawei Technologies Co., Ltd. 2026. All rights reserved.
"""文件解析工具类 — 支持 txt/csv/xlsx/docx/doc 格式

"""

import io
import re
import uuid
from typing import Callable, Dict
from urllib.parse import urlparse

from openjiuwen.core.common.logging import workflow_logger


class FileParser:
    """文件解析工具类

    提供：
    - parse(data, suffix)        — 按扩展名分发解析，返回文本内容
    - get_suffix(file_url)       — 从 URL 提取文件扩展名
    - truncate(text, max_size)  — 按字符数截断
    - build_legal_name(name)    — 文件名校验，非法则用 UUID
    """

    _ILLEGAL_NAME_CHARS = re.compile(r'[\\/:*?"<>|]')

    _UNSUPPORTED_FORMATS: Dict[str, str] = {
        "xls": "不支持 .xls 格式，请转为 .xlsx",
    }

    # --- 内部解析函数 ---

    @staticmethod
    def _decode_bytes(data: bytes) -> str:
        """多编码尝试解码（utf-8/gbk/gb2312/gb18030），兼容中文文件"""
        for enc in ("utf-8", "gbk", "gb2312", "gb18030"):
            try:
                return data.decode(enc)
            except (UnicodeDecodeError, LookupError):
                continue
        return data.decode("utf-8", errors="ignore")

    @staticmethod
    def read_txt(data: bytes) -> str:
        """解析 txt：多编码解码"""
        return FileParser._decode_bytes(data)

    @staticmethod
    def read_csv(data: bytes) -> str:
        """解析 csv：标准库 csv 模块，正确处理引号内逗号"""
        import csv

        text = FileParser._decode_bytes(data)
        lines = []
        for row in csv.reader(io.StringIO(text)):
            lines.append(",".join(row))
        return "\n".join(lines)

    @staticmethod
    def read_xlsx(data: bytes) -> str:
        """解析 xlsx：openpyxl，遍历所有工作表"""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = []
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append("\t".join(cells))
        finally:
            wb.close()
        return "\n".join(lines)

    @staticmethod
    def read_docx(data: bytes) -> str:
        """解析 docx：python-docx，按文档原始顺序遍历段落和表格"""
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph

        doc = Document(io.BytesIO(data))
        parts = []
        for block in doc.element.body:
            if block.tag.endswith("}p"):
                para = Paragraph(block, doc)
                if para.text:
                    parts.append(para.text)
            elif block.tag.endswith("}tbl"):
                table = Table(block, doc)
                for row in table.rows:
                    cells = [cell.text.strip() for cell in row.cells]
                    parts.append("\t".join(cells))
        return "\n".join(parts)

    @staticmethod
    def read_doc(data: bytes) -> str:
        """解析 doc：antiword 命令行工具，写入临时文件后调用

        antiword Windows 版不支持 stdin 输入（'-' 参数），需要用文件路径。
        """
        import os
        import shutil
        import subprocess
        import tempfile

        antiword_path = shutil.which("antiword") or "antiword"
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as f:
            f.write(data)
            tmp_path = f.name

        try:
            result = subprocess.run(
                [antiword_path, "-m", "UTF-8.txt", tmp_path],
                capture_output=True,
            )
            if result.returncode != 0:
                stderr = result.stderr.decode("utf-8", errors="ignore")
                raise ValueError(f"antiword 解析失败: {stderr}")
            return result.stdout.decode("utf-8", errors="ignore")
        finally:
            os.unlink(tmp_path)

    @staticmethod
    def _detect_format(data: bytes) -> str:
        """根据文件头 magic bytes 检测实际格式

        仅检测 ZIP 容器格式（docx/xlsx），OLE2 格式无法区分 doc/xls，留给扩展名判断。
        返回格式名或空字符串（无法识别时回退到扩展名）。
        """
        if len(data) < 4 or data[:4] != b"PK\x03\x04":
            return ""
        import zipfile
        try:
            names = zipfile.ZipFile(io.BytesIO(data)).namelist()
            if "word/document.xml" in names:
                return "docx"
            if "xl/workbook.xml" in names:
                return "xlsx"
        except Exception as e:
            workflow_logger.error("ZIP 格式检测失败: {}", e)
        return ""

    # --- 公开方法 ---

    @classmethod
    def parse(cls, data: bytes, suffix: str) -> str:
        """根据文件内容检测实际格式并分发解析

        优先通过 magic bytes 检测实际格式（docx/xlsx），
        检测不到时回退到扩展名。避免扩展名与实际格式不匹配导致解析失败。
        """
        fmt = cls._detect_format(data) or suffix
        if fmt in cls._UNSUPPORTED_FORMATS:
            return cls._UNSUPPORTED_FORMATS[fmt]
        parser = _PARSERS.get(fmt)
        if parser is None:
            return ""
        return parser(data)

    @classmethod
    def build_legal_name(cls, name: str) -> str:
        """文件名校验：含非法字符或为空则用 UUID（对齐 Java buildLegalName）"""
        if not name or cls._ILLEGAL_NAME_CHARS.search(name):
            return str(uuid.uuid4())
        return name

    @staticmethod
    def get_suffix(file_url: str) -> str:
        """从 URL 提取文件扩展名（小写，不含点）

        只取 URL path 最后一段（文件名）的扩展名，避免域名中的点干扰。
        """
        path = urlparse(file_url).path
        filename = path.rstrip("/").rsplit("/", 1)[-1]
        if "." not in filename:
            return ""
        return filename.rsplit(".", 1)[-1].lower()

    @staticmethod
    def truncate(text: str, max_size: int) -> str:
        """按字符数截断（对齐 Java agent.max-resolve-size，统一字符单位）"""
        if len(text) <= max_size:
            return text
        return text[:max_size]


# 格式→解析函数映射（模块级，避免类外访问受保护成员）
_PARSERS: Dict[str, Callable[[bytes], str]] = {
    "txt": FileParser.read_txt,
    "csv": FileParser.read_csv,
    "xlsx": FileParser.read_xlsx,
    "docx": FileParser.read_docx,
    "doc": FileParser.read_doc,
}
